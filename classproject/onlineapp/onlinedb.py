
import os
import django
from bs4 import BeautifulSoup
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'classproject.settings')
django.setup()
import click
import openpyxl
from onlineapp.models import *


#@click.command()
def main():
    marks_read_excel=openpyxl.load_workbook('marks.xlsx')
    marks_sheet=marks_read_excel.get_sheet_by_name('Sheet')
    rows=marks_sheet.max_row
    for r in range(2, rows+1):
        m = MockTest1()
        nameOfStudent = marks_sheet.cell(row=r,column=1).value
        nameOfStudent = nameOfStudent.split('_')[2]
        m.student = Student.objects.get(db_folder=nameOfStudent)
        m.problem1=marks_sheet.cell(row=r,column=2).value
        m.problem2=marks_sheet.cell(row=r,column=3).value
        m.problem3=marks_sheet.cell(row=r,column=4).value
        m.problem4=marks_sheet.cell(row=r,column=5).value
        m.total=marks_sheet.cell(row=r,column=6).value
        m.save()

    wb = openpyxl.load_workbook('students.xlsx')
    college_sheet = wb['Colleges']
    for i in range(2, college_sheet.max_row + 1):
        c=College()
        c.name=college_sheet.cell(row=i,column=1).value.lower()
        c.acronym = college_sheet.cell(row=i, column=2).value.lower()
        c.location = college_sheet.cell(row=i, column=3).value.lower()
        c.contact = college_sheet.cell(row=i, column=4).value.lower()
        c.save()

    student_sheet=wb['Current']
    for i in range(2,student_sheet.max_row+1):
        s=Student()
        s.name=student_sheet.cell(row=i,column=1).value.lower()
        s.college=College.objects.get(acronym=student_sheet.cell(row=i,column=2).value.lower())
        s.email=student_sheet.cell(row=i,column=3).value.lower()
        s.db_folder=student_sheet.cell(row=i,column=4).value.lower()
        s.save()

if __name__=='__main__':
    main()